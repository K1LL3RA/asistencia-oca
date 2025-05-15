from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import os
import logging
import pandas as pd
from datetime import datetime
from PIL import Image as PilImage

def ajustar_tamano_firma_preciso(ruta_imagen, max_pixels_x=0.0067, max_pixels_y=0.004):
    emu_per_pixel = 9000
    with PilImage.open(ruta_imagen) as img:
        width, height = img.size
        ratio = min(max_pixels_x / width, max_pixels_y / height)
        new_width = int(width * ratio * emu_per_pixel)
        new_height = int(height * ratio * emu_per_pixel)
    return new_width, new_height

def quitar_fondo_blanco(ruta_imagen_original, ruta_salida):
    try:
        img = PilImage.open(ruta_imagen_original).convert("RGBA")
        datas = img.getdata()

        nueva_data = []
        for item in datas:
            if item[0] > 240 and item[1] > 240 and item[2] > 240:
                nueva_data.append((255, 255, 255, 0))
            else:
                nueva_data.append(item)

        img.putdata(nueva_data)
        img.save(ruta_salida, "PNG")
        return ruta_salida
    except Exception as e:
        print(f"Error quitando fondo blanco: {e}")
        return ruta_imagen_original

def insertar_firma_centrada(sheet, fila, columna, ruta_firma, redimension=False):
    try:
        firma = Image(ruta_firma)
        if redimension:
            new_width, new_height = ajustar_tamano_firma_preciso(ruta_firma)
            firma.width = new_width
            firma.height = new_height

        sheet.row_dimensions[fila].height = 35
        col_letra = get_column_letter(columna)
        celda = f"{col_letra}{fila}"
        firma.anchor = celda
        sheet.add_image(firma)
    except Exception as e:
        print(f"Error insertando firma en fila {fila}, columna {columna}: {e}")

def llenar_excel(data, output_path, plantilla_path, texto_g50, texto_g52, ruta_imagen, carpeta_firmas, redimension_firma=False):
    try:
        wb = load_workbook(plantilla_path)
        sheet = wb.active

        def escribir_texto_celda(sheet, celda_ref, texto):
            try:
                is_merged = False
                for merged_range in sheet.merged_cells.ranges:
                    if celda_ref in str(merged_range):
                        min_row = merged_range.min_row
                        min_col = merged_range.min_col
                        cell_to_write = sheet.cell(row=min_row, column=min_col)
                        is_merged = True
                        break
                if not is_merged:
                    cell_to_write = sheet[celda_ref]
                cell_to_write.value = texto
                cell_to_write.alignment = Alignment(horizontal='center', vertical='center')
            except Exception as e:
                logging.error(f"Error escribiendo en celda: {e}")

        capacitador = data[data['Asistió como'].str.lower() == "capacitador"]

        if not capacitador.empty:
            persona = capacitador.iloc[0]
            nombre_capacitador = persona.get('ApellidoNombre', '')
            cargo_capacitador = persona.get('Cargo', '')
            fecha_evento = persona.get('Fecha', '')
            dni_capacitador = str(persona.get('NumeroDocumento', '')).strip()
            sheet['Q44'] = nombre_capacitador
            sheet['Q45'] = cargo_capacitador
            sheet['Q46'] = fecha_evento.strftime('%d/%m/%Y') if isinstance(fecha_evento, pd.Timestamp) else str(fecha_evento)

            if carpeta_firmas:
                extensiones = ['.png', '.jpg', '.jpeg']
                firma_insertada = False
                for ext in extensiones:
                    ruta_firma_original = os.path.join(carpeta_firmas, f"{dni_capacitador}{ext}")
                    if os.path.exists(ruta_firma_original):
                        ruta_firma = quitar_fondo_blanco(ruta_firma_original, ruta_firma_original.replace(ext, "_transparente.png"))
                        insertar_firma_centrada(sheet, 47, 17, ruta_firma, redimension_firma)
                        firma_insertada = True
                        break
                if not firma_insertada:
                    escribir_texto_celda(sheet, "Q47", "Cargado desde SafeSmart")
            else:
                escribir_texto_celda(sheet, "Q47", "Cargado desde SafeSmart")

        participantes = data[data['Asistió como'].str.lower() != "capacitador"]

        for idx, i in enumerate(range(23, 43)):
            if idx < len(participantes):
                row = participantes.iloc[idx]
                sheet[f"C{i}"] = row.get('ApellidoNombre', '')
                sheet[f"K{i}"] = row.get('NumeroDocumento', '')
                sheet[f"N{i}"] = row.get('Área', '')
                sheet[f"Q{i}"] = row.get('Cargo', '')
                dni = str(row.get('NumeroDocumento', '')).strip()

                if carpeta_firmas:
                    extensiones = ['.png', '.jpg', '.jpeg']
                    firma_insertada = False
                    for ext in extensiones:
                        ruta_firma_original = os.path.join(carpeta_firmas, f"{dni}{ext}")
                        if os.path.exists(ruta_firma_original):
                            ruta_firma = quitar_fondo_blanco(ruta_firma_original, ruta_firma_original.replace(ext, "_transparente.png"))
                            insertar_firma_centrada(sheet, i, 24, ruta_firma, redimension_firma)
                            firma_insertada = True
                            break
                    if not firma_insertada:
                        escribir_texto_celda(sheet, f"W{i}", "Cargado desde SafeSmart")
                else:
                    escribir_texto_celda(sheet, f"W{i}", "Cargado desde SafeSmart")

        sheet['G50'] = texto_g50
        sheet['G52'] = texto_g52

        if ruta_imagen and os.path.exists(ruta_imagen):
            img = Image(ruta_imagen)
            img.width = 200
            img.height = 80
            sheet.add_image(img, 'S51')

        if 'Tema Tratado' in data.columns:
            sheet['D18'] = data['Tema Tratado'].mode()[0]
        if 'Análisis del Tema' in data.columns:
            sheet['D19'] = data['Análisis del Tema'].mode()[0]

        if 'Clasificación del Registro' in data.columns:
            clasificaciones = data['Clasificación del Registro'].unique()
            duracion_total = 0
            for clasificacion in clasificaciones:
                if clasificacion in ['Inducción', 'Entrenamiento', 'Capacitación', 'Simulacro de Emergencia', 'Procedimiento', 'Reunión']:
                    duracion_total += 60
                elif clasificacion == 'Charla':
                    duracion_total += 15
                elif clasificacion == 'Difusión':
                    duracion_total += 30
            horas = duracion_total // 60
            minutos = duracion_total % 60
            duracion_texto = f"{horas} hora{'s' if horas > 1 else ''}" if horas else ""
            if minutos:
                duracion_texto += f" {minutos} minuto{'s' if minutos > 1 else ''}"
            sheet['K12'] = duracion_texto.strip()

        if not data.empty:
            participantes_real = data[~data['Asistió como'].str.lower().isin(['capacitador', 'entrenador'])]
            sheet['K11'] = len(participantes_real)


        clasificaciones_tema = {
            'Seguridad': 'Y8',
            'Salud Ocupacional': 'Y9',
            'Medio Ambiente': 'Y10',
            'Calidad': 'Y11',
            'Antisoborno': 'Y12',
        }
        if 'Clasificación del Tema' in data.columns:
            temas = data['Clasificación del Tema'].unique()
            for tema in temas:
                if tema in clasificaciones_tema:
                    sheet[clasificaciones_tema[tema]] = 'X'

        clasificaciones_registro = {
            'Inducción': 'D15',
            'Entrenamiento': 'H15',
            'Charla': 'L15',
            'Simulacro de Emergencia': 'Q15',
            'Capacitación': 'D16',
            'Procedimiento': 'H16',
            'Reunión': 'L16',
            'Difusión': 'Q16',
        }
        if 'Clasificación del Registro' in data.columns:
            tipos = data['Clasificación del Registro'].unique()
            for tipo in tipos:
                if tipo in clasificaciones_registro:
                    sheet[clasificaciones_registro[tipo]] = 'X'

        if 'Fecha' in data.columns:
            fechas = data['Fecha'].unique()
            if len(fechas) > 0:
                fecha_str = fechas[0].strftime('%d/%m/%Y') if isinstance(fechas[0], pd.Timestamp) else str(fechas[0])
                sheet['R8'] = fecha_str
                sheet['R9'] = fecha_str

        hoy = datetime.today().strftime('%d/%m/%Y')
        sheet['G53'] = hoy

        if 'Hora Inicio' in data.columns:
            horas_ini = data['Hora Inicio'].unique()
            if len(horas_ini) > 0:
                sheet['R11'] = str(horas_ini[0])
        if 'Hora Fin' in data.columns:
            horas_fin = data['Hora Fin'].unique()
            if len(horas_fin) > 0:
                sheet['R12'] = str(horas_fin[0])
        if 'Área' in data.columns:
            areas = data['Área'].unique()
            if len(areas) > 0:
                sheet['S16'] = str(areas[0])

        if os.path.exists(output_path):
            try:
                os.remove(output_path)
            except Exception as e:
                logging.error(f"No se pudo eliminar el archivo existente: {e}")
                raise

        try:
            wb.save(output_path)
            print(f"Archivo guardado en {output_path}")
        except PermissionError as e:
            logging.error(f"Permiso denegado al guardar el archivo: {e}")
            raise
        finally:
            wb.close()

    except Exception as e:
        logging.error(f"Error al llenar Excel: {e}")